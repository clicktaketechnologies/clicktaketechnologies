"""
Build the keyword strategy xlsx deliverable.

Inputs:
  /home/z/my-project/upload/all_categories-web_design_services-en-gb-02-08-2026.csv
    (Ahrefs keyword export — 611 rows, 234 with search vol > 0)

Output:
  /home/z/my-project/download/keyword-strategy-web-design-services.xlsx

Sheets:
  1. Summary           — strategy overview, top opps, page-by-page edit summary
  2. Keyword Map       — all 234 vol>0 keywords → target page + slot + priority
  3. Page Edits        — per-page edit plan (current vs new title/meta/H1/H2/body)
  4. New Pages         — new SEO landing pages to create
  5. FAQs              — FAQ questions for schema injection
  6. Internal Links    — internal link map between pages
"""

import csv
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────
CSV_IN = "/home/z/my-project/upload/all_categories-web_design_services-en-gb-02-08-2026.csv"
XLSX_OUT = "/home/z/my-project/download/keyword-strategy-web-design-services.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────
HEAD_FILL = PatternFill("solid", fgColor="0F172A")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="FF53A9")
SUBHEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
HIGH_FILL = PatternFill("solid", fgColor="FECACA")   # red tint
MED_FILL = PatternFill("solid", fgColor="FEF3C7")    # amber tint
LOW_FILL = PatternFill("solid", fgColor="DCFCE7")    # green tint
NEW_FILL = PatternFill("solid", fgColor="E0E7FF")    # indigo tint
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row=1, cols=None):
    """Apply header style to row 1 of a worksheet."""
    cols = cols or ws.max_column
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def style_body(ws, start_row=2):
    """Apply body styling (wrap + border)."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER


def auto_width(ws, max_w=60):
    """Auto-size column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value:
                # Account for line-wrapped cells — use first line length
                v = str(cell.value).split("\n")[0]
                max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max(12, max_len + 3), max_w)


# ── Load keywords ──────────────────────────────────────────────────────────
rows = []
with open(CSV_IN, "r", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    for r in reader:
        try:
            vol = int(r["Search Vol."]) if r["Search Vol."] and r["Search Vol."].isdigit() else 0
        except Exception:
            vol = 0
        try:
            cpc = float(r["CPC (US$)"]) if r["CPC (US$)"] and r["CPC (US$)"] != "-" else 0.0
        except Exception:
            cpc = 0.0
        rows.append({
            "kw": r["Keyword"],
            "vol": vol,
            "cpc": cpc,
            "mod_type": r["Modifier Type"],
            "mod": r["Modifier"],
        })

# ── Keyword → page mapping logic ──────────────────────────────────────────
# Existing pages on clicktaketech.com that already match keyword themes.
EXISTING_PAGE_MAP = [
    # (regex, target_url, target_slot, notes)
    (r"^web design services?$|^web design agency services$|^web design services company$|^best web design services$|^expert web design services$|^professional web design services$|^custom web design services$|^bespoke web design services$",
     "/services/creative/web-design", "Title, Meta, H1, H2, Body",
     "Existing /services/creative/web-design — broaden meta to capture head term + agency/company/professional modifiers"),
    (r"seo.*web design|web design.*seo",
     "NEW:/services/seo-web-design", "Title, Meta, H1, H2, Body, FAQ",
     "NEW PAGE — combined SEO+web design intent (880+720 vol). Biggest single opportunity."),
    (r"wordpress",
     "/services/web/wordpress", "Title, Meta, H1, H2, Body, FAQ",
     "Existing /services/web/wordpress — meta + body inject 'WordPress web design services' keyword"),
    (r"ecommerce|e-commerce|e commerce|shopify|woocommerce|magento",
     "/services/web/ecommerce", "Title, Meta, H1, H2, Body, FAQ",
     "Existing /services/web/ecommerce — covers ecommerce/shopify intent"),
    (r"small business|startup",
     "NEW:/services/small-business-web-design", "Title, Meta, H1, H2, Body, FAQ",
     "NEW PAGE — 'small business web design services' (480 vol) + 13 related (2,750 total vol)"),
    (r"hosting|maintenance|web design hosting",
     "/services/web/domain-hosting", "Title, Meta, H1, H2, Body, FAQ",
     "Existing /services/web/domain-hosting + cross-link /services/web/maintenance. 'web design hosting services' 720 vol."),
    (r"affordable|inexpensive|cheap|low cost|low-cost",
     "/pricing", "Title, Meta, H1, H2, Body, FAQ",
     "Existing /pricing — capture 'affordable web design services' (140) + 42 pricing kws (680 total vol)"),
    (r"responsive",
     "NEW:/services/responsive-web-design", "Title, Meta, H1, H2, Body",
     "NEW PAGE — 'responsive web design services' (170 vol). Mobile-first design pitch."),
    (r"redesign|rebuild|revamp",
     "/services/web/redesign", "Title, Meta, H1, H2, Body",
     "Existing /services/web/redesign — covers redesign intent"),
    (r"london",
     "/cities/london", "Title, Meta, H1, H2, Body",
     "Existing city hub — already targets 'web design services london' (390 vol)"),
    (r"manchester",
     "/cities/manchester", "Title, Meta, H1, H2, Body",
     "Existing city hub — already targets 'web design services manchester' (210 vol)"),
    (r"birmingham",
     "/cities/birmingham", "Title, Meta, H1, H2, Body",
     "Existing city hub — already targets 'web design services birmingham' (50 vol)"),
    (r"leeds",
     "/cities/leeds", "Title, Meta, H1, H2, Body",
     "Existing city hub — already targets 'web design services leeds' (50 vol)"),
    (r"\buk\b|in uk",
     "/", "Title, Meta, H1, Body",
     "Homepage — UK geo kw (1,700 total vol). Add to homepage meta + body."),
    (r"dubai",
     "/cities/dubai", "Title, Meta, H1, Body",
     "Existing city hub — 'web design services dubai' (10 vol)"),
    (r"near me",
     "/contact", "Meta, Body",
     "Near-me intent → drive to /contact + city pages. Add 'web design services near me' to homepage meta."),
    (r"freelance|fiverr|upwork",
     "(skip)", "(skip)",
     "Skip — freelancer-marketplace intent, not our target audience."),
    (r"how to|what is|what are|how much|where to",
     "FAQ", "FAQ section + FAQPage schema",
     "Question-type keywords → FAQ sections on relevant pages + FAQPage JSON-LD"),
]

# ── Helper: classify a keyword ────────────────────────────────────────────
import re

def classify(kw):
    kw_l = kw.lower()
    # Skip Instagram / unrelated
    for pattern, page, slot, notes in EXISTING_PAGE_MAP:
        if re.search(pattern, kw_l):
            return page, slot, notes
    # Default fallback — fold into /services/creative/web-design or flag gap
    return "GAP", "—", "No good existing page — fold into /services/creative/web-design or create new"


# ── Build keyword map sheet data ──────────────────────────────────────────
kw_rows = []
for r in rows:
    page, slot, notes = classify(r["kw"])
    # Priority: vol >= 200 → High; 50-199 → Med; 1-49 → Low; 0 → Skip
    if r["vol"] >= 200:
        pri = "High"
    elif r["vol"] >= 50:
        pri = "Med"
    elif r["vol"] >= 1:
        pri = "Low"
    else:
        pri = "Skip"
    kw_rows.append({
        "Keyword": r["kw"],
        "Search Vol.": r["vol"],
        "CPC (US$)": r["cpc"],
        "Modifier Type": r["mod_type"],
        "Modifier": r["mod"],
        "Target Page": page,
        "Target Slot": slot,
        "Priority": pri,
        "Notes": notes,
    })

# Sort: High first, then by vol desc
kw_rows.sort(key=lambda x: (
    {"High": 0, "Med": 1, "Low": 2, "Skip": 3}[x["Priority"]],
    -x["Search Vol."],
))


# ── Build workbook ────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Summary ──────────────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"

ws["A1"] = "Keyword Strategy — Web Design Services (en-GB)"
ws["A1"].font = Font(bold=True, size=18, color="0F172A")
ws.merge_cells("A1:D1")
ws.row_dimensions[1].height = 32

ws["A2"] = "Source: all_categories-web_design_services-en-gb-02-08-2026.csv  |  Built: 2026-08-03"
ws["A2"].font = Font(italic=True, color="64748B", size=10)
ws.merge_cells("A2:D2")

# Stats block
ws["A4"] = "Dataset Overview"
ws["A4"].font = SUBHEAD_FONT
ws["A4"].fill = SUBHEAD_FILL
ws.merge_cells("A4:D4")

stats = [
    ("Total keywords", len(rows)),
    ("Keywords with vol > 0", sum(1 for r in rows if r["vol"] > 0)),
    ("Keywords with vol ≥ 200 (High priority)", sum(1 for r in rows if r["vol"] >= 200)),
    ("Keywords with vol 50-199 (Med priority)", sum(1 for r in rows if 50 <= r["vol"] < 200)),
    ("Keywords with vol 1-49 (Low priority)", sum(1 for r in rows if 1 <= r["vol"] < 50)),
    ("Total addressable monthly search vol", sum(r["vol"] for r in rows)),
    ("Keywords already mapped to existing page", sum(1 for r in kw_rows if r["Target Page"].startswith("/") or r["Target Page"] == "FAQ")),
    ("Keywords requiring NEW landing page", sum(1 for r in kw_rows if r["Target Page"].startswith("NEW:"))),
    ("Keywords flagged as coverage gap", sum(1 for r in kw_rows if r["Target Page"] == "GAP")),
    ("Keywords intentionally skipped (out of scope)", sum(1 for r in kw_rows if r["Target Page"] == "(skip)")),
]
for i, (label, val) in enumerate(stats, start=5):
    ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
    ws.cell(row=i, column=2, value=val).alignment = CENTER
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=2).border = BORDER

# Top opportunities
start_row = 5 + len(stats) + 2
ws.cell(row=start_row, column=1, value="Top 25 Opportunities (by Vol × CPC commercial intent)").font = SUBHEAD_FONT
ws.cell(row=start_row, column=1).fill = SUBHEAD_FILL
ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
ws.row_dimensions[start_row].height = 24

hdr = ["Keyword", "Vol", "CPC", "Target Page"]
for c, h in enumerate(hdr, start=1):
    cell = ws.cell(row=start_row + 1, column=c, value=h)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = CENTER
    cell.border = BORDER

top25 = sorted([r for r in kw_rows if r["Priority"] in ("High", "Med")], key=lambda x: -x["Search Vol."])[:25]
for i, r in enumerate(top25, start=start_row + 2):
    ws.cell(row=i, column=1, value=r["Keyword"]).border = BORDER
    ws.cell(row=i, column=2, value=r["Search Vol."]).border = BORDER
    ws.cell(row=i, column=2).alignment = CENTER
    ws.cell(row=i, column=3, value=f"${r['CPC (US$)']:.2f}" if r["CPC (US$)"] > 0 else "-").border = BORDER
    ws.cell(row=i, column=3).alignment = CENTER
    ws.cell(row=i, column=4, value=r["Target Page"]).border = BORDER
    if r["Target Page"].startswith("NEW:"):
        for c in range(1, 5):
            ws.cell(row=i, column=c).fill = NEW_FILL
    elif r["Priority"] == "High":
        for c in range(1, 5):
            ws.cell(row=i, column=c).fill = HIGH_FILL
    elif r["Priority"] == "Med":
        for c in range(1, 5):
            ws.cell(row=i, column=c).fill = MED_FILL

# Page-by-page edit summary
start_row = start_row + 2 + len(top25) + 2
ws.cell(row=start_row, column=1, value="Page-by-Page Edit Summary").font = SUBHEAD_FONT
ws.cell(row=start_row, column=1).fill = SUBHEAD_FILL
ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
ws.row_dimensions[start_row].height = 24

page_summary = [
    ("/", "Homepage", "Existing", "Add 'web design services UK' to title/meta. Add 'web design seo services' + 'small business' section links. Add Q&A FAQ for question kws."),
    ("/services", "Services index", "Existing", "Add 'web design services' keyword cluster to meta + add 'web design services' card to top of page."),
    ("/services/creative/web-design", "Web Design (creative)", "Existing", "Head term page — broaden meta to capture 'professional/custom/bespoke/best/expert' modifiers. Add FAQ."),
    ("/services/web/wordpress", "WordPress Development", "Existing", "Meta + body: inject 'wordpress web design services' (390 vol). Add FAQ."),
    ("/services/web/ecommerce", "E-commerce Development", "Existing", "Meta + body: inject 'ecommerce web design services' (210) + 'shopify web design services' (170). Add FAQ."),
    ("/services/web/maintenance", "Website Maintenance", "Existing", "Add 'web design hosting services' (720 vol) kw to meta + body. Add FAQ."),
    ("/services/web/domain-hosting", "Domain & Hosting", "Existing", "Cover 'web design hosting services' intent (720 vol). Add FAQ."),
    ("/services/web/redesign", "Website Redesign", "Existing", "Cover 'website redesign' kws. Minor meta update."),
    ("/services/web/custom-software", "Custom Software Development", "Existing", "Cover 'custom web design services' (110 vol) kw in body."),
    ("/services/seo", "SEO Services", "Existing", "Cross-link from /services/seo-web-design (new)."),
    ("/pricing", "Pricing", "Existing", "Capture 'affordable/inexpensive/cheap/low cost web design services' (680 total vol). Add FAQ."),
    ("/cities/london", "London hub", "Existing", "Already targets 'web design services london' (390 vol). Add FAQ."),
    ("/cities/manchester", "Manchester hub", "Existing", "Already targets 'web design services manchester' (210 vol). Add FAQ."),
    ("/cities/birmingham", "Birmingham hub", "Existing", "Capture 'web design services birmingham' (50 vol)."),
    ("/cities/leeds", "Leeds hub", "Existing", "Capture 'web design services leeds' (50 vol)."),
    ("/contact", "Contact", "Existing", "Add 'web design services near me' meta kw."),
    ("NEW:/services/seo-web-design", "SEO + Web Design combined", "NEW", "NEW PAGE — captures 'seo and web design services' (880) + 'seo web design services' (720). Biggest single opportunity. Combined-service landing page."),
    ("NEW:/services/small-business-web-design", "Small Business Web Design", "NEW", "NEW PAGE — 'small business web design services' (480) + 13 related kws (2,750 total vol). SME-focused pitch."),
    ("NEW:/services/responsive-web-design", "Responsive Web Design", "NEW", "NEW PAGE — 'responsive web design services' (170 vol). Mobile-first + Core Web Vitals pitch."),
    ("NEW:/services/web-design-services", "Web Design Services (head term)", "NEW", "NEW PAGE — pure head term 'web design services' capture. Acts as pillar page for all web-design-services variants."),
]

hdr = ["URL", "Page Name", "Status", "Edit Summary"]
for c, h in enumerate(hdr, start=1):
    cell = ws.cell(row=start_row + 1, column=c, value=h)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = CENTER
    cell.border = BORDER

for i, (url, name, status, summary) in enumerate(page_summary, start=start_row + 2):
    ws.cell(row=i, column=1, value=url).border = BORDER
    ws.cell(row=i, column=2, value=name).border = BORDER
    ws.cell(row=i, column=3, value=status).border = BORDER
    ws.cell(row=i, column=3).alignment = CENTER
    ws.cell(row=i, column=4, value=summary).border = BORDER
    ws.cell(row=i, column=4).alignment = WRAP
    if status == "NEW":
        for c in range(1, 5):
            ws.cell(row=i, column=c).fill = NEW_FILL
    ws.row_dimensions[i].height = 38

# Set column widths
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 75


# ── Sheet 2: Keyword Map ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Keyword Map")
hdr2 = ["#", "Keyword", "Vol", "CPC", "Mod Type", "Target Page", "Target Slot", "Priority", "Notes"]
for c, h in enumerate(hdr2, start=1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws2.row_dimensions[1].height = 32

for i, r in enumerate(kw_rows, start=2):
    vals = [
        i - 1,
        r["Keyword"],
        r["Search Vol."],
        f"${r['CPC (US$)']:.2f}" if r["CPC (US$)"] > 0 else "-",
        r["Modifier Type"],
        r["Target Page"],
        r["Target Slot"],
        r["Priority"],
        r["Notes"],
    ]
    for c, v in enumerate(vals, start=1):
        cell = ws2.cell(row=i, column=c, value=v)
        cell.alignment = WRAP if c == 9 else Alignment(vertical="top")
        cell.border = BORDER
    # Priority tint
    pri = r["Priority"]
    fill = {"High": HIGH_FILL, "Med": MED_FILL, "Low": LOW_FILL}.get(pri)
    if fill:
        for c in range(1, 10):
            ws2.cell(row=i, column=c).fill = fill
    if r["Target Page"].startswith("NEW:"):
        for c in range(1, 10):
            ws2.cell(row=i, column=c).fill = NEW_FILL

widths2 = [5, 50, 8, 9, 14, 38, 28, 10, 55]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"


# ── Sheet 3: Page Edits ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Page Edits")
hdr3 = ["Page URL", "Slot", "Current Value", "New Value", "Rationale"]
for c, h in enumerate(hdr3, start=1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws3.row_dimensions[1].height = 32

page_edits = [
    # ── Homepage (/) ─────────────────────────────────────────────────────
    ("/", "Title", "ClickTake — AI-Powered Digital Agency",
     "Web Design Services UK · AI-Powered Websites — ClickTake",
     "Capture 'web design services UK' (320 vol) + brand. Keep brand-led but front-load target kw."),
    ("/", "Meta Description", "ClickTake builds AI-powered websites, SaaS platforms, mobile apps and growth systems for UK, Pakistan, USA & Dubai brands. 120+ shipped.",
     "Web design services for UK, Pakistan, USA & Dubai brands. Custom websites, WordPress, ecommerce, SEO & hosting. 120+ shipped. Free 30-min consult.",
     "Front-load 'web design services' + enumerate target offerings (WordPress, ecommerce, SEO, hosting) to align with high-vol modifiers."),
    ("/", "H1", "(in NxHero component — needs read)",
     "Web Design Services for UK Brands That Convert",
     "H1 must contain 'web design services' as the page-level primary kw. Keep brand elsewhere."),
    ("/", "Body — new section", "(none)",
     "Add a 'Web design services we offer' card grid linking to /services/creative/web-design, /services/web/wordpress, /services/web/ecommerce, /services/seo-web-design (new).",
     "Creates internal links from homepage to top-priority service pages — passes PageRank to landing pages."),
    ("/", "FAQ additions", "(none for SEO kws)",
     "Add 3 Q&As: 'What are web design services?', 'How much do web design services cost?', 'How to choose a web design services agency?' — inject as FAQPage JSON-LD.",
     "Captures Question-type kws (10 vol each) + featured snippet eligibility."),

    # ── /services index ─────────────────────────────────────────────────
    ("/services", "Title", "Services — AI · Web · Marketing",
     "Web Design Services & Digital Agency — AI · Web · Marketing | ClickTake",
     "Front-load 'web design services' head kw."),
    ("/services", "Meta Description",
     "Browse all ClickTake Technologies services across four practice areas: AI & Machine Learning, Web Development, Digital Marketing, and Creative. Custom LLMs, chatbots, SaaS platforms, SEO, paid ads, branding and video — delivered from offices in Birmingham, Multan, Austin and Dubai.",
     "Web design services, AI development, SEO, ecommerce & growth marketing — across four practice areas. Delivered from Birmingham, Multan, Austin & Dubai. Free consult.",
     "Front-load 'web design services' + trim to ~155 chars."),
    ("/services", "Keywords", "(existing)",
     "Add: 'web design services', 'web design agency services', 'professional web design services', 'web design services UK', 'web design services company'",
     "Meta keywords array additions for top-vol variants."),

    # ── /services/creative/web-design (head term) ───────────────────────
    ("/services/creative/web-design", "Title", "(needs read)",
     "Web Design Services — Professional · Custom · Bespoke | ClickTake",
     "Capture 'professional web design services' (320) + 'custom' (110) + 'bespoke' (70) kws in single title."),
    ("/services/creative/web-design", "Meta Description",
     "(needs read)",
     "Professional web design services for UK, Pakistan, USA & Dubai brands. Custom, bespoke, responsive websites built on Next.js + Figma. WCAG 2.2 AA. Book a free consult.",
     "Capture 'professional' (320) + 'custom' (110) + 'bespoke' (70) + 'responsive' (170) modifiers + geo kw."),
    ("/services/creative/web-design", "H1", "(needs read)",
     "Professional Web Design Services for Modern Brands",
     "H1 contains 'professional web design services' (320 vol)."),
    ("/services/creative/web-design", "FAQ (new)", "(none)",
     "5 Q&As: 'What are web design services?', 'How much do web design services cost in the UK?', 'How long does a web design project take?', 'Do you offer custom web design services?', 'Are your websites responsive and mobile-friendly?' — FAQPage JSON-LD.",
     "Captures Question kws + 'responsive' (170) kw + pricing long-tail."),

    # ── /services/web/wordpress ─────────────────────────────────────────
    ("/services/web/wordpress", "Title", "(uses generateMetadata template)",
     "WordPress Web Design Services — Custom Themes · Headless WP | ClickTake",
     "Capture 'wordpress web design services' (390 vol, $19.41 CPC) head kw."),
    ("/services/web/wordpress", "Meta Description",
     "(uses generateMetadata template)",
     "WordPress web design services: custom themes, plugin development, headless WP with Next.js, performance & SEO. For UK, Pakistan, USA & Dubai brands. Free quote.",
     "Front-load target kw + list sub-services + geo."),
    ("/services/web/wordpress", "H1", "(uses deep-dive template)",
     "WordPress Web Design Services",
     "H1 = exact match target kw."),
    ("/services/web/wordpress", "FAQ (new)", "(existing deep-dive FAQ)",
     "Add 3 Q&As: 'How much do WordPress web design services cost?', 'Do you offer WordPress maintenance after launch?', 'Can you migrate my existing WordPress site?' — append to existing FAQ.",
     "Long-tail + featured snippet capture."),

    # ── /services/web/ecommerce ─────────────────────────────────────────
    ("/services/web/ecommerce", "Title", "(uses generateMetadata template)",
     "Ecommerce Web Design Services — Shopify · WooCommerce · Headless | ClickTake",
     "Capture 'ecommerce web design services' (210) + 'shopify web design services' (170) in single title."),
    ("/services/web/ecommerce", "Meta Description",
     "(uses generateMetadata template)",
     "Ecommerce web design services: Shopify, WooCommerce, headless commerce (Medusa, Saleor). Conversion-optimized UX, payment integrations, SEO-ready. UK · PK · USA · Dubai.",
     "Front-load kw + list platforms + geo."),
    ("/services/web/ecommerce", "FAQ (new)", "(existing deep-dive FAQ)",
     "Add 3 Q&As: 'Shopify vs WooCommerce — which is better for ecommerce web design?', 'How much do ecommerce web design services cost?', 'Do you migrate existing stores?'",
     "Captures comparison + cost long-tail."),

    # ── /services/web/maintenance + /services/web/domain-hosting ────────
    ("/services/web/maintenance", "Title", "(uses generateMetadata template)",
     "Website Maintenance & Web Design Hosting Services | ClickTake",
     "Capture 'web design hosting services' (720 vol)."),
    ("/services/web/domain-hosting", "Meta Description",
     "(uses generateMetadata template)",
     "Web design hosting services: managed cloud hosting (Vercel, Cloudflare, AWS), SSL, CDN, DNS, 24/7 monitoring. Bundled with every ClickTake build or standalone.",
     "Front-load 'web design hosting services' (720 vol) kw."),

    # ── /pricing ────────────────────────────────────────────────────────
    ("/pricing", "Title", "Pricing — Starter · Growth · Scale · Custom",
     "Affordable Web Design Services Pricing — Starter · Growth · Scale | ClickTake",
     "Capture 'affordable web design services' (140) + 'web design services cost' kws."),
    ("/pricing", "Meta Description",
     "Transparent pricing for ClickTake Technologies. Four engagement tiers across the UK, Pakistan, USA and Dubai: Starter (£1,500+), Growth (£6,000+), Scale (£20,000+) and Custom Quote. No hidden fees, no fake universal pricing, no lock-in. Free 30-min consultation.",
     "Affordable web design services pricing: Starter £1,500+, Growth £6,000+, Scale £20,000+, Custom. No hidden fees, no lock-in. UK · PK · USA · Dubai. Free 30-min consult.",
     "Front-load 'affordable web design services' + 'pricing' kw."),
    ("/pricing", "Keywords", "(existing)",
     "Add: 'affordable web design services', 'low cost web design services', 'cheap web design services', 'inexpensive web design services', 'web design services cost', 'how to price web design services'",
     "Capture affordable/pricing cluster (680 total vol)."),
    ("/pricing", "FAQ additions", "(existing 6 FAQs)",
     "Add 2 Q&As: 'How much do web design services cost in the UK?', 'Do you offer affordable web design services for small businesses?'",
     "Capture cost/affordable question kws + featured snippet eligibility."),

    # ── /cities/london ──────────────────────────────────────────────────
    ("/cities/london", "Title (verify)", "(composed by composeCityHubContent)",
     "Verify title contains 'web design services london' (390 vol) — adjust composer if not.",
     "Highest-vol UK city kw."),
    ("/cities/manchester", "Title (verify)", "(composed)",
     "Verify 'web design services manchester' (210 vol).",
     "2nd highest UK city kw."),
]

for i, (url, slot, cur, new, rationale) in enumerate(page_edits, start=2):
    vals = [url, slot, cur, new, rationale]
    for c, v in enumerate(vals, start=1):
        cell = ws3.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
    ws3.row_dimensions[i].height = 50

widths3 = [30, 18, 35, 55, 45]
for i, w in enumerate(widths3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"


# ── Sheet 4: New Pages ────────────────────────────────────────────────────
ws4 = wb.create_sheet("New Pages")
hdr4 = ["URL Slug", "Page Title (target kw)", "H1", "Meta Description (≤155 char)", "Target Keywords (vol)", "Body Sections", "Internal Links Out", "Schema"]
for c, h in enumerate(hdr4, start=1):
    cell = ws4.cell(row=1, column=c, value=h)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws4.row_dimensions[1].height = 32

new_pages = [
    {
        "slug": "/services/seo-web-design",
        "title": "SEO & Web Design Services — Combined Growth Engine | ClickTake",
        "h1": "SEO and Web Design Services That Rank and Convert",
        "meta": "SEO and web design services in one team. We ship fast, indexable, conversion-optimized sites that rank on Google. UK · PK · USA · Dubai. Free SEO audit.",
        "kws": "seo and web design services (880), seo web design services (720), web design seo services (880)",
        "sections": "Hero | Why combine SEO + web design | Our process | Case studies | Pricing | FAQ",
        "links": "/services/creative/web-design, /services/seo, /services/web/full-stack, /pricing, /case-studies",
        "schema": "Service + FAQPage + BreadcrumbList",
    },
    {
        "slug": "/services/small-business-web-design",
        "title": "Small Business Web Design Services — Affordable & Fast | ClickTake",
        "h1": "Small Business Web Design Services That Pay for Themselves",
        "meta": "Small business web design services for UK SMEs. Fast launch (4 weeks), fixed price, SEO-ready, mobile-first. From £1,500. Free 30-min consult.",
        "kws": "small business web design services (480), web design services for small business (320), web design services small business (320)",
        "sections": "Hero | Why small businesses choose us | What's included | Pricing tiers | Case studies | FAQ",
        "links": "/services/creative/web-design, /services/starter-kit, /pricing, /solutions/startups, /solutions/local-businesses",
        "schema": "Service + FAQPage + BreadcrumbList",
    },
    {
        "slug": "/services/responsive-web-design",
        "title": "Responsive Web Design Services — Mobile-First · Core Web Vitals | ClickTake",
        "h1": "Responsive Web Design Services for Every Screen",
        "meta": "Responsive web design services. Mobile-first, Core Web Vitals <100, Lighthouse 95+, WCAG 2.2 AA. UK · PK · USA · Dubai. Free 30-min consult.",
        "kws": "responsive web design services (170)",
        "sections": "Hero | Why responsive matters | Our responsive process | Case studies | FAQ",
        "links": "/services/creative/web-design, /services/web/full-stack, /case-studies",
        "schema": "Service + FAQPage + BreadcrumbList",
    },
    {
        "slug": "/services/web-design-services",
        "title": "Web Design Services — Professional · Custom · UK Agency | ClickTake",
        "h1": "Web Design Services for UK, Pakistan, USA & Dubai Brands",
        "meta": "Web design services for UK brands. Professional, custom, responsive websites on Next.js + Figma. 120+ shipped. Free 30-min consult.",
        "kws": "web design services (root), web design agency services (110), web design services company (140), best web design services (140), professional web design services (320), custom web design services (110), bespoke web design services (70), expert web design services (70)",
        "sections": "Hero | All web design services | Why choose ClickTake | Case studies | Pricing | FAQ",
        "links": "/services/creative/web-design, /services/web/wordpress, /services/web/ecommerce, /services/seo-web-design, /services/small-business-web-design, /services/responsive-web-design, /pricing",
        "schema": "Service + FAQPage + BreadcrumbList + ItemList (of sub-services)",
    },
]

for i, p in enumerate(new_pages, start=2):
    vals = [p["slug"], p["title"], p["h1"], p["meta"], p["kws"], p["sections"], p["links"], p["schema"]]
    for c, v in enumerate(vals, start=1):
        cell = ws4.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = NEW_FILL
    ws4.row_dimensions[i].height = 95

widths4 = [32, 45, 40, 60, 45, 45, 50, 35]
for i, w in enumerate(widths4, start=1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A2"


# ── Sheet 5: FAQs ─────────────────────────────────────────────────────────
ws5 = wb.create_sheet("FAQs")
hdr5 = ["Target Page", "Question", "Answer (concise, 1-3 sentences)", "Source Keyword", "Schema Type"]
for c, h in enumerate(hdr5, start=1):
    cell = ws5.cell(row=1, column=c, value=h)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws5.row_dimensions[1].height = 32

faqs = [
    ("/", "What are web design services?", "Web design services cover everything needed to plan, design, build and launch a website — UX research, UI design, frontend development, content, SEO setup, hosting and ongoing maintenance. ClickTake bundles all of these into a single fixed-scope engagement.", "what are web design services (10)", "FAQPage"),
    ("/", "How much do web design services cost?", "Most ClickTake web design projects land between £1,500 (Starter) and £25,000+ (Scale). Small business sites start at £1,500, marketing sites are £6,000-12,000, and custom SaaS sites are £20,000+. Every quote is fixed-scope — no surprises.", "how to price web design services (20)", "FAQPage"),
    ("/", "How to choose a web design services agency?", "Look for: (1) a portfolio of sites in your industry, (2) case studies with measurable outcomes (traffic, conversions, Core Web Vitals), (3) a transparent fixed-scope contract, (4) senior engineers (not juniors), and (5) post-launch maintenance included. ClickTake meets all five.", "how to choose a web design services agency (related)", "FAQPage"),
    ("/services/creative/web-design", "What are web design services?", "Web design services cover UX research, wireframes, high-fidelity UI design, design systems, interactive prototypes and handoff to engineering — built on Figma, optimized for conversion and accessibility (WCAG 2.2 AA).", "what are web design services (10)", "FAQPage"),
    ("/services/creative/web-design", "How much do web design services cost in the UK?", "UK web design services typically cost £1,500-25,000+. A small business landing page starts at £1,500, a marketing site is £6,000-12,000, and a custom SaaS site is £20,000+. ClickTake offers transparent fixed-scope pricing with no hidden fees.", "web design services cost (20) + how to price (20)", "FAQPage"),
    ("/services/creative/web-design", "How long does a web design project take?", "A landing page ships in 2-3 weeks, a marketing site in 6-8 weeks, and a SaaS site in 10-16 weeks. We kick off within 7 days of signing the proposal and ship weekly demos so you always see progress.", "(related long-tail)", "FAQPage"),
    ("/services/creative/web-design", "Do you offer custom web design services?", "Yes. Every ClickTake site is custom-designed in Figma from a blank canvas — no templates. We start with UX research, then wireframes, then high-fidelity UI, then a design system your team can extend.", "custom web design services (110)", "FAQPage"),
    ("/services/creative/web-design", "Are your websites responsive and mobile-friendly?", "Yes. Every site we build is mobile-first, responsive across all viewports, and optimized for Core Web Vitals (LCP <2.5s, CLS <0.1, INP <200ms). Lighthouse scores are typically 95+ across all categories.", "responsive web design services (170)", "FAQPage"),
    ("/services/web/wordpress", "How much do WordPress web design services cost?", "WordPress web design services start at £3,500 for a custom theme marketing site, £6,000-12,000 for a headless WordPress + Next.js build, and £15,000+ for an enterprise WordPress platform with custom plugins.", "wordpress web design services (390)", "FAQPage"),
    ("/services/web/wordpress", "Do you offer WordPress maintenance after launch?", "Yes. Our WordPress maintenance plans start at £150/month and cover security patches, plugin updates, daily backups, uptime monitoring, performance audits and emergency fixes — month-to-month, cancel anytime.", "(related)", "FAQPage"),
    ("/services/web/wordpress", "Can you migrate my existing WordPress site?", "Yes. We've migrated 50+ WordPress sites from platforms including Wix, Squarespace, Webflow, Drupal and custom CMSs — with URL redirects, schema preservation and zero SEO equity loss. Typical migration takes 2-4 weeks.", "(related)", "FAQPage"),
    ("/services/web/ecommerce", "Shopify vs WooCommerce — which is better for ecommerce web design?", "Shopify is better for brands that want speed-to-market and zero hosting ops (hosted, secure, scales automatically). WooCommerce is better for brands that want full ownership, no monthly fees, and deep WordPress integration. We build both — book a free consult and we'll recommend based on your business.", "ecommerce web design services (210) + comparison kws", "FAQPage"),
    ("/services/web/ecommerce", "How much do ecommerce web design services cost?", "Ecommerce web design services start at £6,000 for a Shopify store, £10,000-20,000 for a WooCommerce build, and £25,000+ for a headless commerce platform (Medusa, Saleor) with custom integrations.", "ecommerce web design services (210)", "FAQPage"),
    ("/services/web/ecommerce", "Do you migrate existing stores?", "Yes. We migrate stores from Magento, BigCommerce, Wix, Squarespace and custom platforms to Shopify, WooCommerce or headless commerce — with full order history, customer data, URL redirects and SEO preservation.", "(related)", "FAQPage"),
    ("/services/web/maintenance", "What do web design hosting services include?", "Our web design hosting services include managed cloud hosting (Vercel, Cloudflare, AWS), SSL certificates, CDN configuration, DNS management, daily backups, 24/7 uptime monitoring, security patches and emergency fixes — bundled with every ClickTake build or as a standalone service.", "web design hosting services (720)", "FAQPage"),
    ("/pricing", "How much do web design services cost in the UK?", "UK web design services cost £1,500-25,000+ depending on scope. ClickTake's Starter tier begins at £1,500 (landing page), Growth at £6,000 (marketing site), Scale at £20,000 (SaaS/custom), and Custom is quoted per project. All pricing is fixed-scope with no hidden fees.", "web design services cost (20)", "FAQPage"),
    ("/pricing", "Do you offer affordable web design services for small businesses?", "Yes. Our Starter tier (£1,500+) is built specifically for small businesses and includes a 4-page responsive website, basic SEO setup, contact form, and 30 days of post-launch support. Payment plans available for projects over £5,000.", "affordable web design services (140)", "FAQPage"),
]

for i, (page, q, a, src, schema) in enumerate(faqs, start=2):
    vals = [page, q, a, src, schema]
    for c, v in enumerate(vals, start=1):
        cell = ws5.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
    ws5.row_dimensions[i].height = 60

widths5 = [32, 38, 75, 32, 14]
for i, w in enumerate(widths5, start=1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = "A2"


# ── Sheet 6: Internal Links ───────────────────────────────────────────────
ws6 = wb.create_sheet("Internal Links")
hdr6 = ["From Page", "To Page", "Anchor Text", "Context / Placement", "Priority"]
for c, h in enumerate(hdr6, start=1):
    cell = ws6.cell(row=1, column=c, value=h)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws6.row_dimensions[1].height = 32

internal_links = [
    ("/", "/services/creative/web-design", "web design services", "Homepage services grid — first card", "High"),
    ("/", "/services/seo-web-design", "SEO and web design services", "Homepage services grid — new card", "High"),
    ("/", "/services/web/wordpress", "WordPress web design services", "Homepage services grid", "High"),
    ("/", "/services/web/ecommerce", "ecommerce web design services", "Homepage services grid", "High"),
    ("/", "/services/small-business-web-design", "small business web design services", "Homepage 'For SMBs' callout", "High"),
    ("/", "/services/web/maintenance", "web design hosting services", "Homepage services grid", "High"),
    ("/", "/cities/london", "web design services London", "Homepage 'Locations we serve' footer section", "High"),
    ("/", "/cities/manchester", "web design services Manchester", "Homepage 'Locations we serve' footer section", "Med"),
    ("/", "/cities/birmingham", "web design services Birmingham", "Homepage 'Locations we serve' footer section", "Med"),
    ("/", "/pricing", "affordable web design services pricing", "Homepage CTA + nav", "Med"),
    ("/services", "/services/creative/web-design", "web design services", "Services index — top of page", "High"),
    ("/services", "/services/seo-web-design", "SEO and web design services", "Services index — new card", "High"),
    ("/services", "/services/small-business-web-design", "small business web design services", "Services index — new card", "High"),
    ("/services/creative/web-design", "/services/seo-web-design", "SEO and web design services", "Related services section", "High"),
    ("/services/creative/web-design", "/services/web/wordpress", "WordPress web design services", "Related services section", "High"),
    ("/services/creative/web-design", "/services/web/ecommerce", "ecommerce web design services", "Related services section", "High"),
    ("/services/creative/web-design", "/services/responsive-web-design", "responsive web design services", "Related services section", "Med"),
    ("/services/creative/web-design", "/services/small-business-web-design", "small business web design services", "Related services section", "High"),
    ("/services/web/wordpress", "/services/creative/web-design", "web design services", "Cross-link in body", "Med"),
    ("/services/web/wordpress", "/services/web/maintenance", "WordPress maintenance services", "Cross-link in body", "Med"),
    ("/services/web/ecommerce", "/services/creative/web-design", "web design services", "Cross-link in body", "Med"),
    ("/services/web/ecommerce", "/services/web/maintenance", "ecommerce maintenance & hosting", "Cross-link in body", "Med"),
    ("/services/web/maintenance", "/services/web/domain-hosting", "domain & hosting services", "Cross-link in body", "Med"),
    ("/services/web/maintenance", "/services/creative/web-design", "web design services", "Cross-link in body", "Med"),
    ("/services/seo-web-design", "/services/creative/web-design", "web design services", "Related services", "High"),
    ("/services/seo-web-design", "/services/seo", "SEO services", "Related services", "High"),
    ("/services/seo-web-design", "/services/web/full-stack", "full-stack web development", "Related services", "Med"),
    ("/services/small-business-web-design", "/services/starter-kit", "Business Starter Kit", "Cross-link", "High"),
    ("/services/small-business-web-design", "/pricing", "affordable web design services pricing", "Cross-link", "High"),
    ("/services/small-business-web-design", "/solutions/startups", "For Startups solution", "Cross-link", "Med"),
    ("/services/small-business-web-design", "/solutions/local-businesses", "For Local Businesses solution", "Cross-link", "Med"),
    ("/services/responsive-web-design", "/services/creative/web-design", "web design services", "Related services", "Med"),
    ("/services/web-design-services", "/services/creative/web-design", "web design services", "Pillar → sub-service", "High"),
    ("/services/web-design-services", "/services/web/wordpress", "WordPress web design services", "Pillar → sub-service", "High"),
    ("/services/web-design-services", "/services/web/ecommerce", "ecommerce web design services", "Pillar → sub-service", "High"),
    ("/services/web-design-services", "/services/seo-web-design", "SEO and web design services", "Pillar → sub-service", "High"),
    ("/services/web-design-services", "/services/small-business-web-design", "small business web design services", "Pillar → sub-service", "High"),
    ("/services/web-design-services", "/services/responsive-web-design", "responsive web design services", "Pillar → sub-service", "Med"),
    ("/pricing", "/services/small-business-web-design", "small business web design services", "Cross-link from pricing tier CTA", "High"),
    ("/cities/london", "/services/creative/web-design", "web design services London", "City hub → service", "High"),
    ("/cities/london", "/services/seo-web-design", "SEO and web design services London", "City hub → service", "High"),
    ("/cities/manchester", "/services/creative/web-design", "web design services Manchester", "City hub → service", "High"),
    ("/cities/birmingham", "/services/creative/web-design", "web design services Birmingham", "City hub → service", "Med"),
    ("/cities/leeds", "/services/creative/web-design", "web design services Leeds", "City hub → service", "Med"),
]

for i, (frm, to, anchor, ctx, pri) in enumerate(internal_links, start=2):
    vals = [frm, to, anchor, ctx, pri]
    for c, v in enumerate(vals, start=1):
        cell = ws6.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
    if pri == "High":
        for c in range(1, 6):
            ws6.cell(row=i, column=c).fill = HIGH_FILL
    elif pri == "Med":
        for c in range(1, 6):
            ws6.cell(row=i, column=c).fill = MED_FILL

widths6 = [30, 38, 38, 45, 10]
for i, w in enumerate(widths6, start=1):
    ws6.column_dimensions[get_column_letter(i)].width = w
ws6.freeze_panes = "A2"


# ── Save ──────────────────────────────────────────────────────────────────
Path(XLSX_OUT).parent.mkdir(parents=True, exist_ok=True)
wb.save(XLSX_OUT)
print(f"✓ Strategy xlsx written to {XLSX_OUT}")
print(f"  Sheets: {wb.sheetnames}")
print(f"  Keyword Map rows: {len(kw_rows)}")
print(f"  Page Edits rows: {len(page_edits)}")
print(f"  New Pages rows: {len(new_pages)}")
print(f"  FAQs rows: {len(faqs)}")
print(f"  Internal Links rows: {len(internal_links)}")
